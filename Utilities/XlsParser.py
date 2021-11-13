import datetime
import os
import openpyxl
from ctypes import ArgumentError
from django.db import transaction
from openpyxl.styles import numbers, PatternFill
import PersonalArea.serializations
from PersonalArea import models
from Utilities import password_generantor


def get_id(record):
    if record is None:
        return generate_id()
    if isinstance(record, int):
        return int(record)
    else:
        if isinstance(record, str):
            if record[0] == '№':
                return int(record[1::])
            else:
                return int(record)
        raise ArgumentError(f"Ячейка заполнена не правильно {record}")


def get_ku(record):
    if isinstance(record, int):
        return record
    else:
        if "дан" in record:
            ku = record.replace('дан', '')
            if isinstance(ku, int):
                return 10 + ku
            else:
                return 10 + len(list(ku))
        elif 'кю' in record:
            return int(record.replace(' ', '').replace("кю", ""))
        else:
            raise ArgumentError(f'Не правильно заполенна ячейка {record}')


def parse_eu_date_to_us(record):
    if len(record.split(' ')) > 1:
        return record.split(' ')[0]
    return datetime.datetime.strptime(record, "%d.%m.%Y").strftime("%Y-%m-%d")


@transaction.atomic
def parseXlcToDb(xlcFile):
    book = openpyxl.open(xlcFile, read_only=True)
    book.iso_dates = True
    sheet = book.active
    fileName = os.path.basename(xlcFile)
    event_name = fileName[0:fileName.find('.')]
    seminars = models.Seminar.objects.filter(name=event_name).values_list('member_id')

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        member_id = get_id(row[4].value)

        if seminars.filter(member_id=member_id).exists():
            continue

        seminar = models.Seminar()
        seminar.name = event_name
        seminar.club = row[7].value
        seminar.trainer = row[8].value
        seminar.city = row[11].value
        seminar.start_date = row[10].value
        seminar.attestation_date = row[12].value
        seminar.oldKu = get_ku(row[3].value)
        seminar.newKu = get_ku(row[13].value)
        seminar.isChild = True if row[14].value == '+' else False
        seminar.examiner = row[15].value

        trainer_id = get_id(row[9].value)
        set_trainer_status(trainer_id)
        if not (models.Aikido_Member.objects.filter(id=member_id).exists()):
            data = {
                'id': int(member_id),
                'password': password_generantor.generate_password(),
                'name': row[1].value,
                'surname': row[0].value,
                'second_name': row[2].value,
                'birthdate': parse_eu_date_to_us(str(row[5].value)),
                'region': int(row[6].value),
                'club': row[7].value,
                'photo': None,
                'isTrainer': False,
                'trainer_id': trainer_id
            }

            serializer = PersonalArea.serializations.RegistrationSerializer(data=data)

            if serializer.is_valid():
                aikiboy = serializer.save()
                seminar.member = aikiboy
                seminar.save()
            else:
                raise ArgumentError(f"Ошибка в строке {row[0].row}")
        else:
            aikiboy = models.Aikido_Member.objects.get(id=member_id)
            seminar.member = aikiboy
            seminar.save()


def set_trainer_status(trainer_id):
    if models.Aikido_Member.objects.filter(id=trainer_id).exists():
        trainer = models.Aikido_Member.objects.get(id=trainer_id)
        trainer.isTrainer = True
        trainer.save()


@transaction.atomic
def createXlsxFromRequests(eventName):
    wb = openpyxl.Workbook(iso_dates=True)
    workSheet = wb.active
    workSheet.append(['Фамилия', 'Имя', 'Отчество', 'степень кю/дан', '#ID', 'Дата рождения',
                      'Регион', 'Клуб', 'Тренер', 'id тренера', 'семинар', 'место проведения', 'аттестация',
                      'Присвоенная степень', 'детский', 'Экзаменатор'])
    set_blue_row(workSheet)
    requests = models.Request.objects.filter(event_name=eventName)
    if not requests.exists():
        raise ArgumentError('Не существует такого мероприятия')

    event = models.Events.objects.get(event_name=eventName)

    for request in requests.values():
        if request['member_id'] is not None:
            member = models.Aikido_Member.objects.get(id=request['member_id'])
            workSheet.append(create_row(request, event, member))
        else:
            workSheet.append(create_row(request, event))

    for i in workSheet.iter_rows(min_row=1, max_row=workSheet.max_row):
        i[5].number_format = numbers.FORMAT_DATE_DDMMYY
        i[10].number_format = numbers.FORMAT_DATE_DDMMYY
        i[12].number_format = numbers.FORMAT_DATE_DDMMYY

    wb.save(f"{eventName}.xlsx")
    wb.close()
    return f"{eventName}.xlsx"


def create_row(request, event, member=None):
    trainer = models.Aikido_Member.objects.get(id=request['trainer_id'])
    oldKu = ''
    if member is not None:
        lastSeminar = models.Seminar.objects.filter(member=member).order_by('attestation_date').values_list('newKu')
        if lastSeminar.exists():
            oldKu = lastSeminar[0][0]
    second_name = '' if request['second_name'] is None else request['second_name']
    member_id = member.id if member is not None else ''
    return [
        request['surname'], request['name'], second_name, oldKu, member_id,
        request['birthdate'],
        trainer.region,
        trainer.club,
        f"{trainer.surname} {trainer.name[0]}.{trainer.second_name[0]}",
        trainer.id,
        event.start_record_date, event.city,
        event.date_of_event, '', '', '']


def set_blue_row(workSheet):
    for row in workSheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = PatternFill(patternType='solid', start_color='538dd5', end_color='538dd5')


def generate_id():
    ids = sorted(models.Aikido_Member.objects.values_list("id",flat=True))
    print(ids)
    limit = ids[-1]
    missingNumbers = list(set(range(1, limit + 1)) - set(ids))
    return missingNumbers[0]

