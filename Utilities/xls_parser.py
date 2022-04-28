from ctypes import ArgumentError
from typing import BinaryIO

import openpyxl
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.db import transaction

import PersonalArea.serializations
from PersonalArea.models import *
from events.models import *
from Utilities import password_generantor
from Utilities import services


@transaction.atomic
def parseXlsToDb(event_slug: str, xlsx_file: InMemoryUploadedFile) -> str:
    book = openpyxl.load_workbook(xlsx_file)
    book.iso_dates = True
    sheet = book.active
    event = services.find_event_by_slug(event_slug)
    if event is None:
        return 'Мероприятие не существует'

    services.delete_achievement_if_exists(event.event_name)

    for row in sheet.iter_rows(min_row=2):
        if row[0].value is None or row[1].value is None or row[5].value is None:
            raise ArgumentError(f"ФИО заполнена не правильно строка-{row[0].row}")

        member_id = services.parse_id_from_xls(row[4].value)
        if Achievements.objects.filter(event_name=event.event_name, member__id=member_id).exists():
            raise ArgumentError(f"Ячейка заполнена не правильно {row[4]}")

        achievement = Achievements()
        achievement.event_name = event.event_name
        achievement.attestation_date = row[12].value
        achievement.received_ku = services.get_ku(row[13].value)
        achievement.is_child = True if row[14].value == '+' else False

        trainer_id = services.parse_id_from_xls(row[9].value)
        services.set_trainer_status(trainer_id)

        if not (Aikido_Member.objects.filter(id=member_id).exists()):
            data = {
                'id': int(member_id),
                'password': password_generantor.generate_password(),
                'name': row[1].value,
                'surname': row[0].value,
                'second_name': row[2].value,
                'birthdate': row[5].value.date(),
                'city': str(row[6].value),
                'club': Club.objects.get(name=row[7].value).id,
                'photo': None,
                'isTrainer': False,
                'trainer_id': trainer_id
            }
            serializer = PersonalArea.serializations.RegistrationSerializer(data=data)
            if serializer.is_valid():
                aikiboy = serializer.save()
                achievement.member = aikiboy
                achievement.save()
            else:
                raise ArgumentError(f"Ошибка в строке {row}")
        else:
            aikiboy = Aikido_Member.objects.get(id=member_id)
            achievement.member = aikiboy
            achievement.save()

    return 'Ведомость загружена'


@transaction.atomic
def createXlsxFromRequests(event_slug: str) -> BinaryIO:
    wb = openpyxl.Workbook(iso_dates=True)
    work_sheet = wb.active
    work_sheet.append(['Фамилия', 'Имя', 'Отчество', 'степень кю/дан', '#ID', 'Дата рождения',
                       'Город', 'Клуб', 'Тренер', 'id тренера', 'семинар', 'место проведения', 'аттестация',
                       'Присвоенная степень', 'детский', 'Экзаменатор'])

    services.set_blue_row(work_sheet)
    event = Events.objects.get(slug=event_slug)

    is_past = True if event.date_of_event <= datetime.date.today() else False

    if is_past:
        records = Achievements.objects.filter(event_name=event.event_name) \
            .prefetch_related('member').prefetch_related('member__club')
        for record in records:
            work_sheet.append(services.create_row_to_past(record, event))

    else:
        requests = Request.objects.filter(event_name=event.id)
        if not requests.exists():
            raise ArgumentError('Заявок на мероприятие нет')

        for request in requests.values():
            member = None
            if request['member_id'] is not None:
                member = Aikido_Member.objects.get(id=request['member_id'])
            work_sheet.append(services.create_row(request, event, member))

    services.set_date_format(work_sheet)

    return services.get_xlsx_file_obj(wb, event_slug)
