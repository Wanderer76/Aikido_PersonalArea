import datetime
import os
import tempfile
from typing import List, Dict, Union, Optional, BinaryIO

from openpyxl import Workbook
from openpyxl.styles import PatternFill, numbers
from openpyxl.worksheet.worksheet import Worksheet

from PersonalArea.models import *
from events.models import *
from ctypes import ArgumentError


def parse_id_from_xls(record: Union[Union[int, str], None]) -> int:
    result = None
    if record is None:
        result = generate_id()
    if isinstance(record, int):
        result = int(record)
    else:
        if isinstance(record, str):
            if record[0] == '№':
                result = int(record.replace(' ', '')[1::])
            else:
                result = int(record)
    if result is None:
        raise ArgumentError(f"Ячейка заполнена не правильно {record}")
    else:
        return result


def get_ku(record: Union[int, Union[str, None]]) -> Optional[int]:
    if isinstance(record, int):
        return record
    else:
        if record == "" or record is None:
            return None
        if 'детский' in record:
            record = record.replace('детский', '')
        if "дан" in record:
            ku = record.replace('дан', '').replace(' ', '')
            if isinstance(ku, int):
                return 10 + ku
            else:
                return 10 + len(list(ku))
        elif 'кю' in record:
            return int(record.replace(' ', '').replace("кю", ""))
        else:
            raise ArgumentError(f'Не правильно заполнена ячейка {record}')


def parse_eu_date_to_us(record: str) -> str:
    if len(record.split(' ')) > 1:
        return record.split(' ')[0]
    return datetime.datetime.strptime(record, "%d.%m.%Y").strftime("%Y-%m-%d")


def set_trainer_status(trainer_id: int) -> None:
    if Aikido_Member.objects.filter(id=trainer_id).exists():
        trainer = Aikido_Member.objects.get(id=trainer_id)
        trainer.isTrainer = True
        trainer.save()
    else:
        raise ArgumentError(f'Тренера с id - {trainer_id} не существует')


def create_row(request: Dict[str, str], event: Events, member=None) -> List[str]:
    trainer = Aikido_Member.objects.get(id=request['trainer_id'])
    last = 0
    ku_suffix = ''
    is_child = False
    if member is not None:
        member_achievements = Achievements.objects.filter(member=member, received_ku__isnull=False).order_by(
            'attestation_date')
        if member_achievements.exists():
            last = member_achievements.values_list('received_ku')[member_achievements.count() - 1][0]
            if member_achievements[member_achievements.count() - 1].is_child:
                ku_suffix = ' детский'
                is_child = True
    second_name = ''
    if request['second_name'] != 'None' and request['second_name'] != '':
        second_name = request['second_name']

    if member is not None and member.second_name is not None:
        second_name = member.second_name

    member_id = member.id if member is not None else ''
    return [
        request['surname'], request['name'], second_name, parse_ku(last) + ku_suffix, member_id,
        request['birthdate'],
        trainer.city,
        trainer.club.name,
        f"{trainer.surname} {trainer.name[0]}.{trainer.second_name[0]}",
        trainer.id,
        event.start_record_date.date(), event.address,
        event.date_of_event, '', '' if not is_child else '+', event.responsible_trainer]


def create_row_to_past(record: Achievements, event: Events):
    member = record.member
    trainer = Aikido_Member.objects.get(id=member.trainer_id)
    ku_suffix = '+' if record.is_child is True else ''
    member_achievements = Achievements.objects.filter(member=member).order_by('attestation_date').values('event_name',
                                                                                                         'received_ku',
                                                                                                         'is_child')
    last_ku = ''
    index = 0
    for i in member_achievements:
        if i['event_name'] == event.event_name:
            break
        index += 1

    if index != 0:
        last_ku = member_achievements[index]['received_ku']

    return [
        member.surname, member.name, member.second_name,
        last_ku,
        member.id,
        member.birthdate,
        member.city,
        member.club.name,
        f"{trainer.surname} {trainer.name[0]}.{trainer.second_name[0]}",
        trainer.id,
        event.start_record_date.date(), event.address,
        event.date_of_event, parse_ku(record.received_ku), ku_suffix, event.responsible_trainer]


def set_blue_row(work_sheet: Worksheet) -> None:
    for row in work_sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.fill = PatternFill(patternType='solid', start_color='538dd5', end_color='538dd5')


def generate_id() -> int:
    ids = sorted(Aikido_Member.objects.values_list("id", flat=True))
    limit = ids[-1]
    missing_numbers = list(set(range(1, limit + 3)) - set(ids))
    return missing_numbers[0]


def delete_achievement_if_exists(event_name: str) -> None:
    achievements = Achievements.objects.filter(event_name=event_name) \
        .prefetch_related("member")
    if achievements.exists():
        hasbiks_id = achievements.values_list("member", flat=True)
        for i in hasbiks_id:
            if Achievements.objects.filter(member_id=i).count() == 1:
                if achievements.filter(member_id=i).exists():
                    Aikido_Member.objects.get(id=i).delete()
            else:
                achievements.get(member_id=i).delete()


def find_event_by_slug(event_slug: str) -> Union[Events, None]:
    return Events.objects.filter(slug=event_slug).first()


def set_date_format(work_sheet: Worksheet) -> None:
    for i in work_sheet.iter_rows(min_row=1, max_row=work_sheet.max_row):
        i[5].number_format = numbers.FORMAT_DATE_DDMMYY
        i[10].number_format = numbers.FORMAT_DATE_DDMMYY
        i[12].number_format = numbers.FORMAT_DATE_DDMMYY


def get_day_before(date_of_event: datetime.date) -> datetime.datetime:
    day_before = date_of_event - datetime.timedelta(1)
    return datetime.datetime(day_before.year, day_before.month, day_before.day, 23, 59, 59, 0)


def parse_ku(value: int) -> str:
    if value > 10:
        return f"{value % 10} дан"
    if value == 0:
        return ""
    return f"{value} кю"


def get_xlsx_file_obj(work_book: Workbook, event_slug: str) -> BinaryIO:
    temp_file = tempfile.NamedTemporaryFile()
    temp_file.name = event_slug + ".xlsx"
    work_book.save(temp_file.name)
    work_book.close()
    stream_file = open(temp_file.name, 'rb')
    original_close = stream_file.close

    def orig_cl():
        original_close()
        os.remove(temp_file.name)

    stream_file.close = orig_cl
    return stream_file
